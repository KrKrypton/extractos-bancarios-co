"""Genera el Excel maestro con el mismo formato que Maestro_Movimientos_Bancarios.xlsx."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

from metadatos import parse_fecha

HEADERS = [
    "Año Gravable",
    "Fecha",
    "Entidad",
    "Número de Cuenta",
    "Producto",
    "Tipo",
    "Monto",
    "Descripción",
    "Origen (Archivo)",
]

COLUMNAS = [
    "anio_gravable",
    "fecha",
    "entidad",
    "numero_cuenta",
    "producto",
    "tipo",
    "monto",
    "descripcion",
    "origen",
]

COLOR_NAVY = "1F3864"
COLOR_GRAY = "595959"
COLOR_LINK = "0563C1"
FILL_HEADER = PatternFill("solid", fgColor=COLOR_NAVY)
FILL_INGRESO = PatternFill("solid", fgColor="C6EFCE")
FILL_EGRESO = PatternFill("solid", fgColor="FFC7CE")
FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_TITLE = Font(bold=True, color=COLOR_NAVY, size=22)
FONT_SUBTITLE = Font(color=COLOR_GRAY, size=10)
FONT_SECTION = Font(bold=True, color=COLOR_NAVY, size=11)
FONT_YEAR_TITLE = Font(bold=True, color=COLOR_NAVY, size=16)
FONT_LINK = Font(color=COLOR_LINK, size=11)
FONT_BODY = Font(size=10)

FMT_FECHA = "dd/mm/yyyy"
FMT_MONTO = '"$ "#,##0.00;[RED]"-$ "#,##0.00'
FMT_MONTO_RESUMEN = '"$ "#,##0;[RED]"-$ "#,##0'

COL_WIDTHS = {"A": 12, "B": 13, "C": 18, "D": 16, "E": 18, "F": 11, "G": 16, "H": 42, "I": 46}


def _cuenta_valor(valor: str | int | float | None):
    if valor is None or valor == "":
        return ""
    s = str(valor).strip()
    if s.isdigit():
        return int(s)
    return s


def _fila_excel(fila: dict) -> list:
    fecha = parse_fecha(str(fila.get("fecha", "")))
    return [
        fila.get("anio_gravable"),
        fecha,
        fila.get("entidad", ""),
        _cuenta_valor(fila.get("numero_cuenta")),
        fila.get("producto", ""),
        fila.get("tipo", ""),
        float(fila.get("monto") or 0),
        fila.get("descripcion", ""),
        fila.get("origen", ""),
    ]


def _aplicar_anchos(ws) -> None:
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width


def _estilo_encabezado_tabla(ws, fila: int, desde: int = 1, hasta: int = 9) -> None:
    for col in range(desde, hasta + 1):
        cell = ws.cell(fila, col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _formato_filas_datos(ws, inicio: int, fin: int) -> None:
    for row in range(inicio, fin + 1):
        c_fecha = ws.cell(row, 2)
        if c_fecha.value:
            c_fecha.number_format = FMT_FECHA
        ws.cell(row, 7).number_format = FMT_MONTO


def _formato_tipo(ws, col: int, inicio: int, fin: int) -> None:
    if fin < inicio:
        return
    letra = get_column_letter(col)
    rango = f"{letra}{inicio}:{letra}{fin}"
    ws.conditional_formatting.add(rango, CellIsRule(operator="equal", formula=['"INGRESO"'], fill=FILL_INGRESO))
    ws.conditional_formatting.add(rango, CellIsRule(operator="equal", formula=['"EGRESO"'], fill=FILL_EGRESO))


def _crear_tabla(ws, nombre: str, ref: str) -> None:
    tabla = Table(displayName=nombre, ref=ref)
    tabla.tableColumns = [TableColumn(id=i + 1, name=h) for i, h in enumerate(HEADERS)]
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabla)


def _hoja_maestro(wb: Workbook, filas: list[dict]) -> None:
    ws = wb.create_sheet("Maestro")
    ws.append(HEADERS)
    _estilo_encabezado_tabla(ws, 1)
    for fila in filas:
        ws.append(_fila_excel(fila))
    fin = len(filas) + 1
    _formato_filas_datos(ws, 2, fin)
    _formato_tipo(ws, 6, 2, fin)
    _aplicar_anchos(ws)
    if filas:
        _crear_tabla(ws, "Maestro_Datos", f"A1:I{fin}")


def _hoja_anio(wb: Workbook, anio: int, filas: list[dict], entidades: list[str]) -> None:
    ws = wb.create_sheet(str(anio))
    nombre_tabla = f"Tabla_{anio}"
    n = len(filas)

    ws.cell(2, 2, f"Resumen año gravable {anio}").font = FONT_YEAR_TITLE
    ws.cell(3, 2, f"{n:,} movimientos".replace(",", "."))
    ws.cell(4, 6, "Por entidad (neto)").font = Font(bold=True, size=10)

    resumen = [
        ("Total Ingresos", f'=SUMIFS({nombre_tabla}[Monto],{nombre_tabla}[Tipo],"INGRESO")'),
        ("Total Egresos", f'=SUMIFS({nombre_tabla}[Monto],{nombre_tabla}[Tipo],"EGRESO")'),
        ("Flujo Neto", "=D5-D6"),
        ("# Transacciones", f"=COUNTA({nombre_tabla}[Tipo])"),
    ]
    for i, (etiqueta, formula) in enumerate(resumen, start=5):
        ws.cell(i, 2, etiqueta).font = Font(bold=True, size=10)
        c = ws.cell(i, 4, formula)
        if i < 8:
            c.number_format = FMT_MONTO_RESUMEN

    for i, entidad in enumerate(entidades):
        row = 5 + i
        ws.cell(row, 6, entidad)
        ws.cell(
            row,
            7,
            f'=SUMIFS({nombre_tabla}[Monto],{nombre_tabla}[Tipo],"INGRESO",{nombre_tabla}[Entidad],"{entidad}")'
            f'-SUMIFS({nombre_tabla}[Monto],{nombre_tabla}[Tipo],"EGRESO",{nombre_tabla}[Entidad],"{entidad}")',
        ).number_format = FMT_MONTO_RESUMEN

    fila_header = 12
    for col, h in enumerate(HEADERS, start=1):
        ws.cell(fila_header, col, h)
    _estilo_encabezado_tabla(ws, fila_header)

    for i, fila in enumerate(filas, start=fila_header + 1):
        for col, val in enumerate(_fila_excel(fila), start=1):
            ws.cell(i, col, val)

    fin = fila_header + n
    if n:
        _formato_filas_datos(ws, fila_header + 1, fin)
        _formato_tipo(ws, 6, fila_header + 1, fin)
        _crear_tabla(ws, nombre_tabla, f"A{fila_header}:I{fin}")

    _aplicar_anchos(ws)


def _hoja_resumen_general(wb: Workbook, anios: list[int], entidades: list[str]) -> None:
    ws = wb.create_sheet("Resumen General")
    ws.cell(2, 2, "Resumen General — Todos los Años Gravables").font = Font(bold=True, size=16, color=COLOR_NAVY)
    ws.cell(3, 2, "Panel dinámico (se actualiza automáticamente con la hoja 'Maestro')").font = FONT_SUBTITLE

    headers_a = ["Año Gravable", "Total Ingresos", "Total Egresos", "Flujo Neto", None, "# Transacciones"]
    for col, h in enumerate(headers_a, start=2):
        if h:
            ws.cell(6, col, h).font = Font(bold=True, size=10)

    headers_b = ["Entidad", "Total Ingresos", "Total Egresos", "Flujo Neto"]
    for col, h in enumerate(headers_b, start=9):
        ws.cell(6, col, h).font = Font(bold=True, size=10)

    first_year_row = 7
    for i, anio in enumerate(anios):
        row = first_year_row + i
        ws.cell(row, 2, anio)
        ws.cell(
            row,
            3,
            f'=SUMIFS(Maestro_Datos[Monto],Maestro_Datos[Tipo],"INGRESO",Maestro_Datos[Año Gravable],B{row})',
        ).number_format = FMT_MONTO_RESUMEN
        ws.cell(
            row,
            4,
            f'=SUMIFS(Maestro_Datos[Monto],Maestro_Datos[Tipo],"EGRESO",Maestro_Datos[Año Gravable],B{row})',
        ).number_format = FMT_MONTO_RESUMEN
        ws.cell(row, 5, f"=C{row}-D{row}").number_format = FMT_MONTO_RESUMEN
        ws.cell(row, 7, f"=COUNTIFS(Maestro_Datos[Año Gravable],B{row})")

    total_row = first_year_row + len(anios)
    ws.cell(total_row, 2, "TOTAL").font = Font(bold=True)
    ws.cell(total_row, 3, f"=SUM(C{first_year_row}:C{total_row - 1})").number_format = FMT_MONTO_RESUMEN
    ws.cell(total_row, 4, f"=SUM(D{first_year_row}:D{total_row - 1})").number_format = FMT_MONTO_RESUMEN
    ws.cell(total_row, 5, f"=SUM(E{first_year_row}:E{total_row - 1})").number_format = FMT_MONTO_RESUMEN
    ws.cell(total_row, 7, f"=SUM(G{first_year_row}:G{total_row - 1})")

    for i, entidad in enumerate(entidades):
        row = first_year_row + i
        ws.cell(row, 9, entidad)
        ws.cell(
            row,
            10,
            f'=SUMIFS(Maestro_Datos[Monto],Maestro_Datos[Tipo],"INGRESO",Maestro_Datos[Entidad],I{row})',
        ).number_format = FMT_MONTO_RESUMEN
        ws.cell(
            row,
            11,
            f'=SUMIFS(Maestro_Datos[Monto],Maestro_Datos[Tipo],"EGRESO",Maestro_Datos[Entidad],I{row})',
        ).number_format = FMT_MONTO_RESUMEN
        ws.cell(row, 12, f"=J{row}-K{row}").number_format = FMT_MONTO_RESUMEN

    matrix_row = total_row + 3
    ws.cell(matrix_row, 2, "Flujo Neto por Año y Entidad").font = FONT_SECTION
    ws.cell(matrix_row + 1, 2, "Año / Entidad").font = Font(bold=True, size=10)
    for col, entidad in enumerate(entidades, start=3):
        ws.cell(matrix_row + 1, col, entidad).font = Font(bold=True, size=10)

    for i, anio in enumerate(anios):
        row = matrix_row + 2 + i
        ws.cell(row, 2, anio)
        for j, entidad in enumerate(entidades):
            col = 3 + j
            col_letra = get_column_letter(col)
            ws.cell(
                row,
                col,
                f'=SUMIFS(Maestro_Datos[Monto],Maestro_Datos[Tipo],"INGRESO",Maestro_Datos[Año Gravable],$B{row},'
                f'Maestro_Datos[Entidad],{col_letra}${matrix_row + 1})'
                f'-SUMIFS(Maestro_Datos[Monto],Maestro_Datos[Tipo],"EGRESO",Maestro_Datos[Año Gravable],$B{row},'
                f'Maestro_Datos[Entidad],{col_letra}${matrix_row + 1})',
            ).number_format = FMT_MONTO_RESUMEN

    if anios:
        last_year_row = first_year_row + len(anios) - 1
        chart = BarChart()
        chart.type = "col"
        chart.title = "Ingresos vs Egresos por Año Gravable"
        chart.y_axis.numFmt = FMT_MONTO_RESUMEN
        chart.width = 18
        chart.height = 10
        cats = Reference(ws, min_col=2, min_row=first_year_row, max_row=last_year_row)
        ingresos = Reference(ws, min_col=3, min_row=6, max_row=last_year_row)
        egresos = Reference(ws, min_col=4, min_row=6, max_row=last_year_row)
        chart.add_data(ingresos, titles_from_data=True)
        chart.add_data(egresos, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "B20")

    if entidades:
        pie = PieChart()
        pie.title = "Egresos Totales por Entidad"
        first_ent_row = first_year_row
        last_ent_row = first_year_row + len(entidades) - 1
        labels = Reference(ws, min_col=9, min_row=first_ent_row, max_row=last_ent_row)
        values = Reference(ws, min_col=11, min_row=6, max_row=last_ent_row)
        pie.add_data(values, titles_from_data=True)
        pie.set_categories(labels)
        pie.width = 14
        pie.height = 10
        ws.add_chart(pie, "J20")

    _aplicar_anchos(ws)
    ws.column_dimensions["L"].width = 16


def _hoja_inicio(wb: Workbook, total: int, anios: list[int]) -> None:
    ws = wb.create_sheet("Inicio")
    ws.column_dimensions["B"].width = 90

    rango = f"{anios[0]} – {anios[-1]}" if anios else "—"
    hoy = datetime.now().strftime("%d/%m/%Y")

    ws.cell(2, 2, "Maestro de Movimientos Bancarios").font = FONT_TITLE
    ws.cell(
        3,
        2,
        f"Información gravable {rango}  ·  {total:,} movimientos  ·  Generado el {hoy}".replace(",", "."),
    ).font = FONT_SUBTITLE

    secciones = [
        (6, "¿Qué contiene este archivo?", None),
        (7, "• Hoja 'Resumen General': panel con totales por año, por entidad y gráficos comparativos.", FONT_BODY),
        (
            8,
            "• Hojas por año: el detalle de movimientos de cada año gravable, con su propio resumen arriba.",
            FONT_BODY,
        ),
        (
            9,
            "• Hoja 'Maestro': la base de datos completa (todos los años juntos), lista para tablas dinámicas.",
            FONT_BODY,
        ),
        (11, "Filtros", None),
        (12, "• Todas las hojas de datos están convertidas en Tabla de Excel con filtros en cada encabezado.", FONT_BODY),
        (13, "• Para limpiar todos los filtros: pestaña Datos → Borrar.", FONT_BODY),
        (15, "Tablas dinámicas (Pivot Tables)", None),
        (
            16,
            "• Resúmenes en 'Resumen General' y en cada año se actualizan solos con fórmulas SUMIFS.",
            FONT_BODY,
        ),
        (
            17,
            "• Para tablas dinámicas interactivas: Insertar → Tabla dinámica → origen hoja 'Maestro'.",
            FONT_BODY,
        ),
        (19, "Buenas prácticas", None),
        (20, "• Verde = Ingreso, Rojo = Egreso (aplica en todas las hojas de detalle).", FONT_BODY),
        (21, "• No borres ni muevas los encabezados de las tablas.", FONT_BODY),
        (23, "Contenido del archivo", None),
        (24, "→  Resumen General", FONT_LINK),
        (25, "→  Maestro", FONT_LINK),
    ]
    for row, texto, font in secciones:
        cell = ws.cell(row, 2, texto)
        cell.font = font or FONT_SECTION

    fila = 26
    for anio in anios:
        ws.cell(fila, 2, f"→  {anio}").font = FONT_LINK
        fila += 1


def generar_maestro_workbook(filas: list[dict], ruta: Path) -> None:
    anios = sorted({f["anio_gravable"] for f in filas if f.get("anio_gravable")})
    entidades = sorted({f["entidad"] for f in filas if f.get("entidad")})
    por_anio: dict[int, list[dict]] = {a: [] for a in anios}
    for fila in filas:
        anio = fila.get("anio_gravable")
        if anio in por_anio:
            por_anio[anio].append(fila)

    wb = Workbook()
    wb.remove(wb.active)

    _hoja_maestro(wb, filas)
    for anio in anios:
        _hoja_anio(wb, anio, por_anio[anio], entidades)
    _hoja_resumen_general(wb, anios, entidades)
    _hoja_inicio(wb, len(filas), anios)

    orden = ["Inicio", "Resumen General", *[str(a) for a in anios], "Maestro"]
    wb._sheets.sort(key=lambda ws: orden.index(ws.title))

    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)
